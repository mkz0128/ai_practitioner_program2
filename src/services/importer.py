from pathlib import Path
from typing import Any, BinaryIO, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from src.domain.models import Dataset, Order, Package, Priority, Vehicle, VehicleStatus, Zone
from src.services.errors import FieldError, ValidationReport

SHEET_FIELDS: dict[str, tuple[str, ...]] = {
    "orders": (
        "order_id",
        "zone_code",
        "city",
        "district",
        "location_label",
        "latitude",
        "longitude",
        "time_slot",
        "declared_package_count",
        "priority",
        "note",
    ),
    "packages": ("package_id", "order_id", "weight_kg"),
    "vehicles": (
        "vehicle_id",
        "vehicle_name",
        "max_load_kg",
        "current_load_kg",
        "service_zone_codes",
        "depot_id",
        "status",
        "note",
    ),
    "zones": (
        "zone_code",
        "zone_name",
        "covered_cities",
        "covered_districts",
        "center_latitude",
        "center_longitude",
        "tdx_city_codes",
        "adjacent_zone_codes",
        "enabled",
    ),
}

LIST_FIELDS = {
    "service_zone_codes",
    "covered_cities",
    "covered_districts",
    "tdx_city_codes",
    "adjacent_zone_codes",
}


def _safe_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text[:80] if text else None


def _split_list(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    return tuple(part.strip() for part in str(value).split("|") if part.strip())


def _coerce_bool(value: Any) -> Any:
    """Normalize common spreadsheet boolean encodings without weakening strict models."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().upper()
        if normalized in {"TRUE", "YES", "Y", "1"}:
            return True
        if normalized in {"FALSE", "NO", "N", "0"}:
            return False
    return value


def _rows(sheet: Any, sheet_name: str, errors: list[FieldError]) -> list[dict[str, Any]]:
    values = list(sheet.values)
    if not values:
        errors.append(FieldError(path=sheet_name, code="SHEET_EMPTY", message="工作表不得為空。"))
        return []
    expected = list(SHEET_FIELDS[sheet_name])
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    if headers != expected:
        errors.append(
            FieldError(
                path=f"{sheet_name}[1]",
                code="INVALID_HEADERS",
                message="欄位順序或名稱不符合固定契約。",
            )
        )
        return []
    result: list[dict[str, Any]] = []
    for _row_number, row in enumerate(values[1:], start=2):
        if all(value is None for value in row):
            continue
        record: dict[str, Any] = {}
        for index, field in enumerate(expected):
            value = row[index] if index < len(row) else None
            record[field] = _split_list(value) if field in LIST_FIELDS else value
        result.append(record)
    return result


def parse_workbook(
    source: str | Path | BinaryIO, source_filename: str = "workbook.xlsx"
) -> tuple[Dataset | None, ValidationReport]:
    errors: list[FieldError] = []
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception:
        return None, ValidationReport(
            is_valid=False,
            errors=[FieldError(path="file", code="INVALID_XLSX", message="無法讀取 .xlsx 檔案。")],
        )
    expected_sheets = tuple(SHEET_FIELDS)
    if tuple(workbook.sheetnames) != expected_sheets:
        errors.append(
            FieldError(
                path="workbook",
                code="INVALID_SHEETS",
                message="必須包含且只包含 orders、packages、vehicles、zones 四張工作表。",
            )
        )
    if errors:
        return None, ValidationReport(is_valid=False, errors=errors)
    records = {name: _rows(workbook[name], name, errors) for name in expected_sheets}
    if errors:
        return None, ValidationReport(is_valid=False, errors=errors)

    orders: list[Order] = []
    packages: list[Package] = []
    vehicles: list[Vehicle] = []
    zones: list[Zone] = []
    for sheet_name, model_type, target in (
        ("orders", Order, orders),
        ("packages", Package, packages),
        ("vehicles", Vehicle, vehicles),
        ("zones", Zone, zones),
    ):
        for row_number, record in enumerate(records[sheet_name], start=2):
            try:
                normalized = dict(record)
                if sheet_name == "orders":
                    raw_priority = normalized.get("priority") or Priority.NORMAL.value
                    normalized["priority"] = Priority(str(raw_priority).strip().upper())
                    item = model_type.model_validate(normalized)
                elif sheet_name == "vehicles":
                    raw_status = normalized.get("status")
                    normalized["status"] = VehicleStatus(str(raw_status).strip().upper())
                    item = model_type.model_validate(normalized)
                elif sheet_name == "zones":
                    normalized["enabled"] = _coerce_bool(normalized.get("enabled"))
                    item = model_type.model_validate(normalized)
                else:
                    item = model_type.model_validate(normalized)
                cast(list[Any], target).append(item)
            except Exception:
                errors.append(
                    FieldError(
                        path=f"{sheet_name}[{row_number}]",
                        code="FIELD_VALIDATION_ERROR",
                        message="欄位型別、必填值或範圍不符合規則。",
                    )
                )
    if errors:
        return None, ValidationReport(is_valid=False, errors=errors)
    package_map: dict[str, list[Package]] = {order.order_id: [] for order in orders}
    for package in packages:
        package_map.setdefault(package.order_id, []).append(package)
    orders = [
        order.model_copy(update={"packages": tuple(package_map.get(order.order_id, []))})
        for order in orders
    ]
    dataset = Dataset(
        orders=tuple(orders),
        packages=tuple(packages),
        vehicles=tuple(vehicles),
        zones=tuple(zones),
        source_filename=source_filename,
    )
    report = validate_dataset(dataset)
    return dataset, report


def validate_dataset(dataset: Dataset) -> ValidationReport:
    errors: list[FieldError] = []
    for label, values, field in (
        ("orders", dataset.orders, "order_id"),
        ("packages", dataset.packages, "package_id"),
        ("vehicles", dataset.vehicles, "vehicle_id"),
        ("zones", dataset.zones, "zone_code"),
    ):
        seen: set[str] = set()
        for value in values:
            identifier = getattr(value, field)
            if identifier in seen:
                errors.append(
                    FieldError(
                        path=f"{label}.{field}",
                        code="DUPLICATE_ID",
                        message=f"{identifier} 重複。",
                        value_summary=identifier,
                    )
                )
            seen.add(identifier)
    order_ids = {order.order_id for order in dataset.orders}
    package_by_order: dict[str, list[Package]] = {order.order_id: [] for order in dataset.orders}
    for package in dataset.packages:
        if package.order_id not in order_ids:
            errors.append(
                FieldError(
                    path=f"packages.{package.package_id}.order_id",
                    code="ORPHAN_PACKAGE",
                    message="package 指向不存在的 order。",
                    value_summary=package.order_id,
                )
            )
        else:
            package_by_order[package.order_id].append(package)
    zone_map = {zone.zone_code: zone for zone in dataset.zones}
    for order in dataset.orders:
        packages = package_by_order[order.order_id]
        if not packages:
            errors.append(
                FieldError(
                    path=f"orders.{order.order_id}.packages",
                    code="ORDER_WITHOUT_PACKAGE",
                    message="每張訂單至少需要一件 package。",
                    value_summary=order.order_id,
                )
            )
        if len(packages) != order.declared_package_count:
            errors.append(
                FieldError(
                    path=f"orders.{order.order_id}.declared_package_count",
                    code="PACKAGE_COUNT_MISMATCH",
                    message="宣告件數與實際 packages 不一致。",
                    value_summary=str(order.declared_package_count),
                )
            )
        if order.zone_code not in zone_map or not zone_map[order.zone_code].enabled:
            errors.append(
                FieldError(
                    path=f"orders.{order.order_id}.zone_code",
                    code="UNKNOWN_ZONE",
                    message="配送區域不存在或未啟用。",
                    value_summary=order.zone_code,
                )
            )
        elif (
            order.city not in zone_map[order.zone_code].covered_cities
            or order.district not in zone_map[order.zone_code].covered_districts
        ):
            errors.append(
                FieldError(
                    path=f"orders.{order.order_id}.district",
                    code="ZONE_MEMBERSHIP_ERROR",
                    message="城市/行政區不屬於宣告的營運區域。",
                    value_summary=order.district,
                )
            )
    valid_vehicle_zones = set(zone_map)
    for vehicle in dataset.vehicles:
        unknown = [zone for zone in vehicle.service_zone_codes if zone not in valid_vehicle_zones]
        if unknown:
            errors.append(
                FieldError(
                    path=f"vehicles.{vehicle.vehicle_id}.service_zone_codes",
                    code="UNKNOWN_SERVICE_ZONE",
                    message="車輛服務區域不存在。",
                    value_summary="|".join(unknown),
                )
            )
        if vehicle.current_load_kg > vehicle.max_load_kg:
            errors.append(
                FieldError(
                    path=f"vehicles.{vehicle.vehicle_id}.current_load_kg",
                    code="INVALID_CURRENT_LOAD",
                    message="目前載重不可超過最大載重。",
                    value_summary=str(vehicle.current_load_kg),
                )
            )
    return ValidationReport(is_valid=not errors, errors=errors)
