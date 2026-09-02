from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.api.main import app, store
from src.domain.models import Dataset
from src.services.importer import parse_workbook
from src.services.matrix import SimulatedRouteProvider
from src.services.planner import build_baseline
from src.services.validator import validate_plan

SAMPLE_WORKBOOK = Path(__file__).parents[1] / "data" / "samples" / "demo-delivery-40-orders.xlsx"
client = TestClient(app)


def _dataset() -> Dataset:
    dataset, report = parse_workbook(SAMPLE_WORKBOOK)
    assert report.is_valid and dataset is not None
    return dataset


def test_z4_112kg_is_split_beyond_100kg_vehicle_and_validated() -> None:
    dataset = _dataset()
    matrix = SimulatedRouteProvider().build(dataset)
    plan = build_baseline(dataset, matrix)
    validation = validate_plan(dataset, plan, matrix)

    assert validation.valid, validation.model_dump()
    z4_ids = {order.order_id for order in dataset.orders if order.zone_code == "Z4"}
    z4_assignments = {
        order_id: route.vehicle_id
        for route in plan.routes
        for order_id in route.order_ids
        if order_id in z4_ids
    }
    assert sum(order.total_weight_kg for order in dataset.orders if order.zone_code == "Z4") == 112
    assert "VEH-003" in set(z4_assignments.values())
    assert set(z4_assignments.values()) != {"VEH-002"}
    assert all(route.planned_load_kg <= route.max_load_kg for route in plan.routes)


def test_missing_required_cells_report_order_package_field_and_manual_review(
    tmp_path: Path,
) -> None:
    workbook = load_workbook(SAMPLE_WORKBOOK)
    orders = workbook["orders"]
    order_headers = {cell.value: index for index, cell in enumerate(orders[1], start=1)}
    orders.cell(row=2, column=order_headers["location_label"]).value = None
    orders.cell(row=3, column=order_headers["time_slot"]).value = None
    packages = workbook["packages"]
    package_headers = {cell.value: index for index, cell in enumerate(packages[1], start=1)}
    packages.cell(row=2, column=package_headers["weight_kg"]).value = None
    path = tmp_path / "missing-fields.xlsx"
    workbook.save(path)

    dataset, report = parse_workbook(path)

    assert dataset is None
    assert report.is_valid is False
    assert report.requires_manual_review is True
    errors = {error.path: error for error in report.errors}
    assert "orders.ORD-001.location_label" in errors
    assert "orders.ORD-002.time_slot" in errors
    assert "packages.PKG-001-01.weight_kg" in errors
    assert all(error.code == "MISSING_REQUIRED_FIELD" for error in errors.values())
    assert all(error.requires_manual_review for error in errors.values())


def test_time_window_conflict_is_explicit_and_plan_remains_validator_safe() -> None:
    dataset = _dataset()
    target = dataset.orders[0].model_copy(update={"latitude": 90.0, "longitude": 0.0})
    changed_orders = (target, *dataset.orders[1:])
    conflict_dataset = dataset.model_copy(update={"orders": changed_orders})
    matrix = SimulatedRouteProvider().build(conflict_dataset)
    plan = build_baseline(conflict_dataset, matrix)
    validation = validate_plan(conflict_dataset, plan, matrix)

    assert plan.unassigned_reasons[target.order_id] == "TIME_WINDOW_CONFLICT"
    assert validation.valid, validation.model_dump()


def test_order_over_all_eligible_capacity_is_unassignable_not_silent() -> None:
    dataset = _dataset()
    heavy_package = dataset.packages[0].model_copy(update={"weight_kg": 1000.0})
    heavy_order = dataset.orders[0].model_copy(update={"packages": (heavy_package,)})
    changed_orders = (heavy_order, *dataset.orders[1:])
    changed_packages = (heavy_package, *dataset.packages[1:])
    heavy_dataset = dataset.model_copy(
        update={"orders": changed_orders, "packages": changed_packages}
    )
    matrix = SimulatedRouteProvider().build(heavy_dataset)
    plan = build_baseline(heavy_dataset, matrix)
    validation = validate_plan(heavy_dataset, plan, matrix)

    assert plan.unassigned_reasons[heavy_order.order_id] == "UNASSIGNABLE"
    assert heavy_order.order_id in plan.unassigned_orders
    assert validation.valid, validation.model_dump()


def test_plan_api_reasons_are_deterministic_evidence_for_every_assigned_order() -> None:
    store.datasets.clear()
    store.plans.clear()
    store.current_versions.clear()
    with SAMPLE_WORKBOOK.open("rb") as workbook:
        imported = client.post(
            "/api/v1/datasets/import-excel",
            files={
                "file": (
                    SAMPLE_WORKBOOK.name,
                    workbook,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert imported.status_code == 201, imported.text
    plan_response = client.post(
        "/api/v1/plans",
        json={
            "dataset_id": imported.json()["dataset_id"],
            "algorithm": "BASELINE",
            "route_provider_preference": "SIMULATED",
        },
    )
    assert plan_response.status_code == 201, plan_response.text
    body = plan_response.json()
    assigned_stops = [
        stop for vehicle in body["vehicles"] for stop in vehicle["stops"]
    ]
    assert assigned_stops
    for stop in assigned_stops:
        reason = stop["reason"]
        evidence = reason["evidence"]
        assert reason["summary"]
        assert evidence["vehicle_zone_eligible"] is True
        assert evidence["order_weight_kg"] > 0
        assert evidence["post_assignment_load_kg"] >= evidence["order_weight_kg"]
        assert 0 < evidence["post_assignment_utilization"] <= 1
        assert evidence["time_window_legal"] is True
        assert evidence["leg_distance_m"] == stop["leg_distance_m"]
        assert evidence["leg_duration_s"] == stop["leg_duration_s"]
        assert evidence["distance_basis"] == "fixed_simulated_matrix"
